#!/usr/bin/env python3
# -*- coding: utf-8 -*--
import io
import warnings
from typing import List, Dict, Tuple, Optional

import pandas as pd
import requests
import streamlit as st
import time
from datetime import datetime
from urllib.parse import quote, urlparse, parse_qs, urlencode, urlunparse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== tijdzone helpers =====
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo

session = requests.Session()

def now_naive_in_tz(tz_str: str) -> pd.Timestamp:
    return pd.Timestamp(datetime.now(ZoneInfo(tz_str))).tz_localize(None)

def now_aware_in_tz(tz_str: str) -> pd.Timestamp:
    return pd.Timestamp(datetime.now(ZoneInfo(tz_str)))

# ===== versie =====
__version__ = "3.1.1"

# ===== warnings onderdrukken (macOS LibreSSL/urllib3) =====
warnings.filterwarnings(
    "ignore",
    message=r"urllib3 v2 only supports OpenSSL.*",
    category=Warning,
    module=r"urllib3\.__init__"
)

# === VASTE INSTELLINGEN ===
debug_fetch = False
DEFAULT_CLIENT_ID = "K662D1WXrt"
TZ = "Europe/Amsterdam"
DAYS_AHEAD = 60
WEEK_OFFSET = -1
FIELDS = "naam,datumvanaf,datumtot,tijdvanaf,tijdtot,lokatie,heledag"

BAR_CODES = ["701", "741", "761"]
CK_CODES  = ["442"]
WEEK_LABEL = "short"          # of "iso"
SAT_ONLY_CK = True            # CommissieKamer alleen zaterdag
sportlink_stats = {
    "calls": 0,
    "retries": 0,
    "failures": 0
}

# ===== Verenigingsactiviteiten =====
ACTIVITIES_DAYS_AHEAD = 60

ACTIVITIES_FIELDS = (
    "kalendernaam,kalendersoort,activiteit,datumvan,datumtm,"
    "heledag,beheerders,opmerkingen,plaats,url"
)

def build_activities_url(days: int, client_id: str, fields: str = ACTIVITIES_FIELDS) -> str:
    base = "https://data.sportlink.com/verenigingsactiviteiten"
    url = f"{base}?aantaldagen={int(days)}&client_id={client_id}"
    if fields:
        url += f"&fields={quote(fields)}"
    return url

# Wedstrijden (programma)
PROGRAM_DAYS_AHEAD = 60
PROGRAM_FIELDS = ("wedstrijddatum,wedstrijdnummer,thuisteamclubrelatiecode,"
                  "uitteamclubrelatiecode,thuisteam,uitteam,competitiesoort,aanvangstijd")
CKC_CLUBRELATIECODE = "BBDZ08H"

DAYS_NL = ["Maandag","Dinsdag","Woensdag","Donderdag","Vrijdag","Zaterdag","Zondag"]
DAY_COLORS = {
    "Maandag":"FFDDEBF7","Dinsdag":"FFE2EFDA","Woensdag":"FFFFF2CC",
    "Donderdag":"FFFCE4D6","Vrijdag":"FFE7E6E6","Zaterdag":"FFE4DFEC","Zondag":"FFF8CBAD"
}
WEEK_COLORS = [
    "D9E1F2",  # licht blauw
    "E2EFDA",  # licht groen
    "FCE4D6",  # licht oranje
    "EAD1DC",  # licht paars
    "FFF2CC",  # licht geel
]

# Shifts per dag: (Tijd-van, Tijd-tot)
DEFAULT_SLOTS: Dict[str, List[Tuple[str, str]]] = {
    "Maandag":   [("18:00","19:00"), ("19:00","20:00"), ("20:00","22:30")],
    "Dinsdag":   [("18:00","19:00"), ("19:00","20:00"), ("20:00","22:30")],
    "Woensdag":  [("17:00","18:00"), ("18:00","19:00"), ("19:00","22:00")],
    "Donderdag": [("18:00","19:00"), ("19:00","20:00"), ("20:00","22:30")],
    "Vrijdag":   [("18:00","20:30"), ("20:30","23:00")],
    "Zaterdag":  [("07:30","10:00"), ("10:00","12:30"),
                  ("12:30","15:00"), ("15:00","17:30"),
                  ("17:30","20:00"), ("20:00","22:30")],
    "Zondag":    [("10:00","12:30"), ("12:30","15:00")],
}

# Dropbox handmatige input
DROPBOX_INPUT_URL = "https://www.dropbox.com/scl/fi/ukcs87y9h1j27uyzcotig/rooster_input.txt?rlkey=fx0ayzshabo7zikun620m61hh&st=vtrlzr8k&dl=0"

DROPBOX_OVERRIDE_URL = "https://www.dropbox.com/scl/fi/w1711x6bzna5lniz0cvkw/Afgeschermd.txt?rlkey=cy3ltl3j427eqtg3k9ylwvc01&st=e6z7qa2n&dl=0"

# ---------- helpers ----------
# ------ build activities calendar -------
def build_activities_calendar_matrix(df_activities: pd.DataFrame):

    if df_activities.empty:
        return pd.DataFrame()

    df = df_activities.copy()
    df = df.sort_values(["ISO_Year", "Week", "Datum", "Tijd"])

    weeks = sorted({
        (int(y), int(w))
        for y, w in zip(df["ISO_Year"], df["Week"])
    })

    rows = []

    grouped = df.groupby(["ISO_Year", "Week", "Dag"])

    for (y, w) in weeks:

        # 🔹 HEADER RIJ (datums per dag)
        header_row = [""]  # kolom A leeg

        # 🔹 DATA RIJ (activiteiten)
        data_row = [f"{y}-W{w:02d}"]

        for dag in DAYS_NL:

            group = grouped.get_group((y, w, dag)) if (y, w, dag) in grouped.groups else None

            if group is not None:
                date = group.iloc[0]["Datum"]
                header_row.append(f"{dag} ({date.strftime('%d-%b')})")

                lines = []
                for _, r in group.sort_values("Tijd").iterrows():
                    tijd = r["Tijd"]
                    naam = r["Activiteit"]

                    if r.get("IsAllDay", False):
                        lines.append(f"{naam}")
                    else:
                        lines.append(f"{tijd} {naam}")

                data_row.append("\n".join(lines))

            else:
                header_row.append(dag)
                data_row.append("")

        rows.append(header_row)
        rows.append(data_row)

    columns = ["Week"] + DAYS_NL
    return pd.DataFrame(rows, columns=columns)

def format_activities_calendar_sheet(ws, matrix: pd.DataFrame, tz_str: str):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    bold = Font(bold=True)
    grey = Font(color="888888")

    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    WEEK_COLORS = ["D9E1F2", "E2EFDA", "FCE4D6", "EAD1DC", "FFF2CC"]

    # 🔹 kolombreedtes
    ws.column_dimensions['A'].width = 14
    for col in ws.iter_cols(min_col=2):
        ws.column_dimensions[col[0].column_letter].width = 26

    # 🔹 rijen verwerken (per week 2 rijen)
    for r_idx in range(2, ws.max_row + 1):

        row = ws[r_idx]
        is_header_row = (r_idx % 2 == 0)  # eerste rij van weekblok

        week_index = (r_idx - 2) // 2
        color = WEEK_COLORS[week_index % len(WEEK_COLORS)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for c_idx, cell in enumerate(row):

            cell.border = border

            if is_header_row:
                # 🔹 header rij (dag + datum)
                cell.fill = fill
                cell.font = bold
                cell.alignment = center

            else:
                # 🔹 data rij
                cell.alignment = wrap

                if c_idx == 0:
                    cell.font = bold
                    cell.alignment = center

        # 🔹 rijhoogtes
        if is_header_row:
            ws.row_dimensions[r_idx].height = 28
        else:
            ws.row_dimensions[r_idx].height = 95

    # 🔹 timestamp (A1)
    now = now_naive_in_tz(tz_str)
    cell = ws.cell(row=1, column=1)
    cell.value = now.strftime("%d-%m %H:%M")
    cell.font = grey
    
    
def load_afgeschermd_overrides_from_dropbox(debug=False):
    overrides = {}
    warnings = []

    try:
        url = _ensure_dropbox_direct(DROPBOX_OVERRIDE_URL)
        txt = session.get(url, timeout=30).text

        for lineno, line in enumerate(txt.splitlines(), start=1):
            raw = line.strip()

            if not raw or raw.startswith("#"):
                continue

            parts = raw.split()

            # basis check
            if len(parts) < 4:
                warnings.append(f"regel {lineno}: te weinig velden → '{raw}'")
                continue

            date_str, time_str, loc = parts[0], parts[1], parts[2].lower()
            name = " ".join(parts[3:]).strip()

            # datum validatie
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
            except Exception:
                warnings.append(f"regel {lineno}: ongeldige datum → '{raw}'")
                continue

            # tijd validatie
            try:
                datetime.strptime(time_str, "%H:%M")
            except Exception:
                warnings.append(f"regel {lineno}: ongeldige tijd → '{raw}'")
                continue

            # locatie validatie
            if loc not in ("bar", "ck"):
                warnings.append(f"regel {lineno}: onbekende locatie '{loc}' → '{raw}'")
                continue

            # naam check
            if not name:
                warnings.append(f"regel {lineno}: ontbrekende naam → '{raw}'")
                continue

            key = (date_str, time_str, loc)
            overrides.setdefault(key, []).append(name)

        # logging
        if debug:
            st.write(f"Overrides geladen: {len(overrides)} geldige sleutels")

        if warnings:
            st.warning(
                "⚠️ Ongeldige override regels overgeslagen:\n\n- " +
                "\n- ".join(warnings)
            )

    except Exception as e:
        st.error(f"Fout bij laden overrides: {e}")

    return overrides

def apply_afgeschermd_overrides(matrix: pd.DataFrame,
                                 overrides: dict,
                                 location: str,
                                 week_label_style: str,
                                 debug: bool = False):

    applied_count = 0
    added_count = 0

    for (dag, van, tot, regel) in matrix.index:

        # Alleen Namen-regels met echte tijdsloten
        if regel != "Namen" or not van:
            continue

        for col in matrix.columns:

            cell = matrix.loc[(dag, van, tot, regel), col]

            # Week/jaar bepalen
            try:
                if week_label_style == "iso":
                    y, w = map(int, col.split("-W"))
                else:
                    w = int(col.split()[1])
                    y = now_naive_in_tz(TZ).isocalendar().year
            except Exception:
                continue

            # Datum reconstrueren
            try:
                monday = pd.Timestamp.fromisocalendar(y, w, 1)
                date = (monday + pd.Timedelta(days=DAYS_NL.index(dag))).strftime("%Y-%m-%d")
            except Exception:
                continue

            key = (date, van, location)

            # Geen override → skip
            if key not in overrides:
                continue

            names_override = overrides[key]

            # Huidige celinhoud
            names_existing = cell.split("\n") if cell else []

            # Zoek Afgeschermd
            afgeschermd_idx = [
                i for i, n in enumerate(names_existing)
                if n.strip().lower() == "afgeschermd"
            ]

            n = len(afgeschermd_idx)
            m = len(names_override)

            # ===== CASE 1: vervangen =====
            if n > 0:
                replace_count = min(n, m)

                for i in range(replace_count):
                    idx = afgeschermd_idx[i]
                    names_existing[idx] = names_override[i]

                applied_count += replace_count

                if debug:
                    st.write(f"[override] {location} {date} {van}: {replace_count}/{n} vervangen")

            # ===== CASE 2: toevoegen =====
            elif m > 0:
                # voorkom dubbele namen
                existing_set = {n.strip().lower() for n in names_existing if n.strip()}

                to_add = [n for n in names_override if n.strip().lower() not in existing_set]

                if to_add:
                    names_existing.extend(to_add)
                    added_count += len(to_add)

                    if debug:
                        st.write(f"[override] {location} {date} {van}: {len(to_add)} toegevoegd")

            # Terugschrijven (alleen als iets veranderd is)
            if n > 0 or (m > 0 and to_add):
                matrix.loc[(dag, van, tot, regel), col] = "\n".join(names_existing)

    if debug:
        st.info(f"Overrides toegepast: {applied_count} vervangen, {added_count} toegevoegd")

def month_short_nl(m:int) -> str:
    return ["jan","feb","mrt","apr","mei","jun","jul","aug","sept","okt","nov","dec"][m-1]

def build_urls(taskcodes: List[str], days: int, client_id: str,
               weekoffset: int = -1, fields: Optional[str] = FIELDS) -> List[str]:
    base = "https://data.sportlink.com/vrijwilligers"
    urls = []
    for code in taskcodes:
        url = (f"{base}?vrijwilligerstaakcode={code}"
               f"&aantaldagen={int(days)}&client_id={client_id}&weekoffset={int(weekoffset)}")
        if fields:
            url += f"&fields={quote(fields)}"
        urls.append(url)
    return urls

def build_program_url(days: int, client_id: str, fields: str = PROGRAM_FIELDS,
                      eigenwedstrijden: str = "JA", thuis: str = "JA", uit: str = "NEE",
                      gebruiklokaleteamgegevens: str = "NEE") -> str:
    base = "https://data.sportlink.com/programma"
    url = (f"{base}?aantaldagen={int(days)}&client_id={client_id}"
           f"&eigenwedstrijden={eigenwedstrijden}&thuis={thuis}&uit={uit}"
           f"&gebruiklokaleteamgegevens={gebruiklokaleteamgegevens}")
    if fields:
        url += f"&fields={quote(fields)}"
    return url

def http_get_json(url: str):

    headers = {
        "User-Agent": "CKC-Rooster/Streamlit",
        "Accept": "application/json"
    }

    # taakcode of endpoint uit URL halen
    task_code = None
    endpoint = "unknown"

    try:
        parsed = urlparse(url)

        if "vrijwilligers" in parsed.path:
            endpoint = "vrijwilligers"
            qs = parse_qs(parsed.query)
            task_code = qs.get("vrijwilligerstaakcode", ["?"])[0]

        elif "programma" in parsed.path:
            endpoint = "programma"

    except Exception:
        pass

    max_retries = 4
    sportlink_stats["calls"] += 1

    for attempt in range(1, max_retries + 1):

        try:

            if debug_fetch:

                if endpoint == "vrijwilligers":
                    st.write(f"Fetching vrijwilligers code {task_code} (poging {attempt})")

                else:
                    st.write(f"Fetching {endpoint} (poging {attempt})")

            r = session.get(
                url,
                timeout=(10, 30),
                headers=headers
            )

            r.raise_for_status()

            if not r.content:
                if debug_fetch:
                    st.warning(f"⚠️ Lege response ({endpoint} {task_code})")
                return []

            data = r.json()

            if isinstance(data, dict):
                data = data.get("items", [])

            if not isinstance(data, list):
                return []

            if debug_fetch:
                if endpoint == "vrijwilligers":
                    st.write(f"✔ {len(data)} records ontvangen voor code {task_code}")
                else:
                    st.write(f"✔ {len(data)} wedstrijden ontvangen")

            return data

        except (
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout,
            requests.exceptions.ChunkedEncodingError,
            requests.exceptions.ContentDecodingError,
            requests.exceptions.HTTPError,
            ValueError
        ) as e:

            if debug_fetch:

                if endpoint == "vrijwilligers":
                    st.warning(
                        f"⚠️ Netwerkfout code {task_code} ({attempt}/{max_retries})"
                    )
                else:
                    st.warning(
                        f"⚠️ Netwerkfout {endpoint} ({attempt}/{max_retries})"
                    )

            if attempt < max_retries:
                sportlink_stats["retries"] += 1
                time.sleep(1.5)
            else:

                if debug_fetch:

                    if endpoint == "vrijwilligers":
                        st.error(f"❌ Ophalen mislukt voor vrijwilligers code {task_code}")

                    else:
                        st.error(f"❌ Ophalen mislukt voor {endpoint}")
                        
                    sportlink_stats["failures"] += 1

                return []

from concurrent.futures import ThreadPoolExecutor
from itertools import chain

@st.cache_data(ttl=300)
def fetch_all(urls: List[str], debug: bool = False):

    results = []

    with ThreadPoolExecutor(max_workers=10) as exe:
        fetched = list(exe.map(http_get_json, urls))

    for url, r in zip(urls, fetched):

        task_code = None
        endpoint = "onbekend"

        try:
            parsed = urlparse(url)

            if "vrijwilligers" in parsed.path:
                endpoint = "vrijwilligers"
                qs = parse_qs(parsed.query)
                task_code = qs.get("vrijwilligerstaakcode", ["?"])[0]

        except Exception:
            pass

        if debug:

            if endpoint == "vrijwilligers":
                st.write(f"✔ vrijwilligers code {task_code}: {len(r)} records")
            else:
                st.write(f"✔ {endpoint}: {len(r)} records")

        results.append(r if isinstance(r, list) else [])

    return list(chain.from_iterable(results))
    
def _pick(colnames, candidates):
    for c in candidates:
        if c in colnames: return c
    lower = {c.lower(): c for c in colnames}
    for c in candidates:
        if c.lower() in lower: return lower[c.lower()]
    return None

def normalize_dataframe(data, tz_str: str):
    df_raw = pd.DataFrame(data)
    cols = df_raw.columns.tolist()

    c_naam = _pick(cols, ["Naam","naam","Vrijwilliger","vrijwilliger","vrijwilligerNaam","displayName"])
    c_dv   = _pick(cols, ["Datum vanaf","datumvanaf","start","Start","DatumVanaf","startDatumTijd","startDateTime"])
    c_dt   = _pick(cols, ["Datum tot","datumtot","eind","Eind","DatumTot","eindDatumTijd","endDateTime"])
    c_tv   = _pick(cols, ["Tijd vanaf","tijdvanaf","startTijd","starttijd","StartTijd"])
    c_tt   = _pick(cols, ["Tijd tot","tijdtot","eindtijd","EindTijd","endTijd","TijdTot"])

    df = df_raw.copy()
    if c_tv is None and c_dv is not None:
        tmp = pd.to_datetime(df[c_dv], errors="coerce", utc=True)
        df["Tijd vanaf"] = tmp.dt.tz_convert(tz_str).dt.strftime("%H:%M"); c_tv = "Tijd vanaf"
    if c_tt is None and c_dt is not None:
        tmp2 = pd.to_datetime(df[c_dt], errors="coerce", utc=True)
        df["Tijd tot"] = tmp2.dt.tz_convert(tz_str).dt.strftime("%H:%M"); c_tt = "Tijd tot"

    out = pd.DataFrame({
        "Naam": df[c_naam] if c_naam else "",
        "Datum vanaf": df[c_dv] if c_dv else "",
        "Datum tot": df[c_dt] if c_dt else "",
        "Tijd vanaf": df[c_tv] if c_tv else "",
        "Tijd tot": df[c_tt] if c_tt else "",
    })

    dat = pd.to_datetime(out["Datum vanaf"], errors="coerce", utc=True)
    dat = dat.dt.tz_convert(tz_str).dt.tz_localize(None)
    out["Datum"] = dat
    out = out.dropna(subset=["Datum"])

    out["Week"] = out["Datum"].dt.isocalendar().week.astype(int)
    out["ISO_Year"] = out["Datum"].dt.isocalendar().year.astype(int)
    out["Weekdag_num"] = out["Datum"].dt.weekday
    out["Dag"] = out["Weekdag_num"].map(lambda i: DAYS_NL[i] if pd.notna(i) else None)

    t_from = pd.to_datetime(out["Tijd vanaf"], format="%H:%M", errors="coerce")
    out["Tijd vanaf"] = t_from.dt.strftime("%H:%M").fillna("")
    t_to = pd.to_datetime(out["Tijd tot"], format="%H:%M", errors="coerce")
    out["Tijd tot"] = t_to.dt.strftime("%H:%M").fillna("")
    return out

def filter_from_current_week(df: pd.DataFrame, tz_str: str) -> pd.DataFrame:
    if "Datum" not in df.columns:
        return df.iloc[0:0].copy()
    if not pd.api.types.is_datetime64_any_dtype(df["Datum"]):
        df = df.copy()
        df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
    df = df.dropna(subset=["Datum"]).copy()
    if df.empty:
        return df
    now_naive = now_naive_in_tz(tz_str)
    monday_naive = (now_naive - pd.Timedelta(days=int(now_naive.weekday()))).normalize()
    return df[df["Datum"].dt.normalize() >= monday_naive].copy()

def monday_of_week(d: pd.Timestamp) -> pd.Timestamp:
    return d - pd.Timedelta(days=int(d.weekday()))

def derive_weeks(df: pd.DataFrame, tz_str: str, horizon_weeks_if_empty=4):
    if not df.empty:
        weeks_pairs = sorted({(int(y), int(w)) for y, w in zip(df["ISO_Year"], df["Week"])})
        week_mondays = {}
        for (y, w) in weeks_pairs:
            d0 = df.loc[(df["ISO_Year"] == y) & (df["Week"] == w), "Datum"].iloc[0]
            week_mondays[(y, w)] = monday_of_week(d0)
        return weeks_pairs, week_mondays
    now = now_naive_in_tz(tz_str)
    mon0 = now - pd.Timedelta(days=int(now.weekday()))
    weeks_pairs = []; week_mondays = {}
    for i in range(horizon_weeks_if_empty):
        mon = (mon0 + pd.Timedelta(days=7*i)).normalize()
        iso = mon.isocalendar(); pair = (int(iso.year), int(iso.week))
        weeks_pairs.append(pair); week_mondays[pair] = mon
    return weeks_pairs, week_mondays

# ===== matrix met subregels (intern) =====
REGELS = ["Handmatig", "Wedstrijden", "Namen"]

def build_empty_matrix(slots: Dict[str, List[Tuple[str,str]]],
                       tz_str: str,
                       days_subset=None,
                       horizon_weeks_if_empty=4,
                       week_label_style: str="short",
                       weeks_pairs=None, week_mondays=None) -> pd.DataFrame:
    if weeks_pairs is None or week_mondays is None:
        raise ValueError("weeks_pairs en week_mondays zijn verplicht")

    def week_label(pair: Tuple[int,int]) -> str:
        y, w = pair
        return f"{y}-W{w:02d}" if week_label_style == "iso" else f"Week {w}"

    days_to_use = DAYS_NL if days_subset is None else [d for d in DAYS_NL if d in days_subset]
    rows = []
    for d in days_to_use:
        rows.append((d, "", "", ""))  # dag-header
        for (van, tot) in slots.get(d, []):
            for r in REGELS:
                rows.append((d, van, tot, r))

    matrix = pd.DataFrame(
        "",
        index=pd.MultiIndex.from_tuples(rows, names=["Dag","Tijd-van","Tijd-tot","Regel"]),
        columns=[week_label(p) for p in weeks_pairs]
    )

    # Kolom-headers met datum
    for p in weeks_pairs:
        mon = week_mondays[p]; col = week_label(p)
        for d in days_to_use:
            day_date = (mon + pd.Timedelta(days=DAYS_NL.index(d))).strftime("%d-%b")
            matrix.loc[(d,"","", ""), col] = f"{d} ({day_date})"

    return matrix

# ===== helpers voor tijd =====
def _hhmm_to_minutes(hhmm: str) -> int:
    try:
        h, m = hhmm.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return -1
        
# == v2.11.0: default en custom slots samenvoegen  
def merge_custom_slots_into_defaults(
    df_list,
    base_slots: Dict[str, List[Tuple[str, str]]],
    activities_df: Optional[pd.DataFrame] = None
) -> Tuple[Dict[str, List[Tuple[str, str]]], List[str]]:    
    """
    Haalt custom diensten uit data en voegt ze toe aan DEFAULT_SLOTS.
    Inclusief:
    - chronologisch sorteren
    - overlap corrigeren (aansluiten)
    - deduplicatie
    Retourneert: (nieuwe_slots, warnings)
    """
    warnings = []
    slots = {d: list(v) for d, v in base_slots.items()}
    
    base_set = {
        (d, sv, st)
        for d, lst in base_slots.items()
        for (sv, st) in lst
    }

    # verzamel alle diensten
    for df in df_list:
        if df is None or df.empty:
            continue

        for _, r in df.iterrows():
            d = r.get("Dag")
            t_from = str(r.get("Tijd vanaf") or "").strip()
            t_to   = str(r.get("Tijd tot") or "").strip()

            if not d or not t_from or not t_to:
                continue

            if d not in slots:
                slots[d] = []

            # 🔹 NIEUW: skip als exact bestaand slot
            if (d, t_from, t_to) in base_set:
                continue
            
            new_from = _hhmm_to_minutes(t_from)
            new_to   = _hhmm_to_minutes(t_to)

            if new_from < 0 or new_to <= new_from:
                continue

            adjusted_from = new_from
            adjusted_to   = new_to

            # overlap met bestaande slots corrigeren
            for (sv, st) in slots[d]:
                sv_m = _hhmm_to_minutes(sv)
                st_m = _hhmm_to_minutes(st)

                # overlap → inkorten
                if max(adjusted_from, sv_m) < min(adjusted_to, st_m):
                    if adjusted_from < sv_m:
                        adjusted_to = min(adjusted_to, sv_m)
                    else:
                        adjusted_from = max(adjusted_from, st_m)

            if adjusted_to <= adjusted_from:
                warnings.append(f"{d} {t_from}-{t_to} volledig overlapt met bestaand slot")
                continue

            new_slot = (
                f"{adjusted_from//60:02d}:{adjusted_from%60:02d}",
                f"{adjusted_to//60:02d}:{adjusted_to%60:02d}"
            )

            if new_slot not in slots[d]:
                slots[d].append(new_slot)
            
                if (adjusted_from != new_from) or (adjusted_to != new_to):
                    warnings.append(f"{d} {t_from}-{t_to} aangepast naar {new_slot[0]}-{new_slot[1]}")
                else:
                    warnings.append(f"{d} {t_from}-{t_to} toegevoegd als nieuw tijdslot")
                    
    # 🔥 Activiteiten als extra tijdsloten
    if activities_df is not None and not activities_df.empty:
        for _, r in activities_df.iterrows():
            d = r.get("Dag")
            t = str(r.get("Tijd") or "").strip()

            if not d or not t:
                continue

            if d not in slots:
                slots[d] = []

            start = _hhmm_to_minutes(t)
            end = start + 30

            # 🔴 check overlap met bestaande slots
            overlap = False
            for (sv, st) in slots[d]:
                sv_m = _hhmm_to_minutes(sv)
                st_m = _hhmm_to_minutes(st)
            
                if max(start, sv_m) < min(end, st_m):
                    overlap = True
                    break
            
            if overlap:
                continue  # skip activiteit slot
            
            new_slot = (
                f"{start//60:02d}:{start%60:02d}",
                f"{end//60:02d}:{end%60:02d}"
            )

            if new_slot not in slots[d]:
                slots[d].append(new_slot)
                warnings.append(f"📅 Activiteit slot toegevoegd: {d} {new_slot[0]}-{new_slot[1]}")
    
    
    # sorteren + aansluiten
    for d in slots:
        day_slots = sorted(slots[d], key=lambda x: _hhmm_to_minutes(x[0]))

        merged = []
        for sv, st in day_slots:
            if not merged:
                merged.append((sv, st))
                continue

            prev_sv, prev_st = merged[-1]
            prev_end = _hhmm_to_minutes(prev_st)
            cur_start = _hhmm_to_minutes(sv)

            # aansluiten
            if cur_start < prev_end:
                sv = prev_st  # schuif start op
            
            # 🔴 NIEUW: check of slot nog geldig is
            if _hhmm_to_minutes(st) <= _hhmm_to_minutes(sv):
                continue  # skip invalid slot
            
            merged.append((sv, st))
            
        slots[d] = merged
        warnings = list(dict.fromkeys(warnings))

    return slots, warnings


# ===== v2.8: namen in ALLE overlappende slots plaatsen + WARNINGS (met datum) =====
def fill_names(matrix: pd.DataFrame, df: pd.DataFrame,
               slots: Dict[str, List[Tuple[str,str]]],
               week_label_style: str) -> List[str]:
    """
    Plaats namen in ALLE slots waarmee de dienst overlapt.
    Retourneert waarschuwingen (incl. datum) voor diensten die niet geplaatst konden worden.
    """
    warnings_list: List[str] = []

    for _, r in df.iterrows():
        d = r.get("Dag")
        t_from = str(r.get("Tijd vanaf") or "").strip()
        t_to   = str(r.get("Tijd tot")   or "").strip()
        naam   = str(r.get("Naam") or "").strip()

        # Datum-string (YYYY-MM-DD) voor meldingen
        date_obj = r.get("Datum")
        date_str = ""
        try:
            if isinstance(date_obj, pd.Timestamp):
                date_str = date_obj.strftime("%Y-%m-%d")
        except Exception:
            pass

        if not d or not t_from:
            continue

        try:
            w = int(r["Week"]); y = int(r["ISO_Year"])
        except Exception:
            continue
        col = (f"{y}-W{w:02d}" if week_label_style == "iso" else f"Week {w}")
        if col not in matrix.columns:
            continue

        rf = _hhmm_to_minutes(t_from)
        rt = _hhmm_to_minutes(t_to) if t_to else -1

        overlapped = []
        for (sv, st) in slots.get(d, []):
            sv_m = _hhmm_to_minutes(sv); st_m = _hhmm_to_minutes(st)
            if rf < 0 or sv_m < 0 or st_m <= sv_m:
                continue
            if rt > rf:  # eindtijd aanwezig → overlap
                if max(rf, sv_m) < min(rt, st_m):
                    overlapped.append((sv, st))
            else:
                # geen eindtijd → containment op start
                if sv_m <= rf < st_m:
                    overlapped.append((sv, st))

        if not overlapped:
            beschrijving = f"{date_str} ({d}) {t_from}-{t_to}" if t_to else f"{date_str} ({d}) {t_from}"
            msg = f"{naam or '(naam onbekend)'} — {beschrijving} kon niet worden geplaatst"
            warnings_list.append(msg)
            continue

        if not naam:
            naam = "(naam onbekend)"

        for (sv, st) in overlapped:
            key = (d, sv, st, "Namen")
            if key not in matrix.index:
                beschrijving = f"{date_str} ({d}) {t_from}-{t_to}" if t_to else f"{date_str} ({d}) {t_from}"
                warnings_list.append(f"{naam} — {beschrijving} past niet in de roosterstructuur")
                continue
            cur = matrix.loc[key, col]
            matrix.loc[key, col] = (cur + "\n" + naam) if cur else naam

    return warnings_list

# === fill_manual (met bestaan-check) ===
def fill_manual(matrix: pd.DataFrame, annotations, slots: Dict[str, List[Tuple[str,str]]],
                week_label_style: str):
    def week_label(y, w): return f"{y}-W{w:02d}" if week_label_style=="iso" else f"Week {w}"

    for a in annotations:
        label = week_label(a["iso_year"], a["iso_week"])
        if label not in matrix.columns:
            continue

        tot = None
        for (v, t) in slots.get(a["day"], []):
            if v == a["time_from"]:
                tot = t; break
        if tot is None:
            continue

        key = (a["day"], a["time_from"], tot, "Handmatig")
        if key not in matrix.index:
            continue  # voorkomt KeyError bij dagen die niet bestaan (bv. CK: alleen Zaterdag)

        cur = matrix.loc[key, label]
        txt = a["text"].strip()
        if txt:
            matrix.loc[key, label] = (cur + "\n" + txt) if cur else txt

def build_match_index_for_overlap(df_program: pd.DataFrame) -> Dict[tuple, List[Tuple[int, str]]]:
    idx: Dict[tuple, List[Tuple[int, str]]] = {}
    for _, r in df_program.iterrows():
        y, w, d = int(r["ISO_Year"]), int(r["Week"]), r["Dag"]
        try:
            h, m = str(r["Tijd"]).split(":"); tmin = int(h)*60+int(m)
        except Exception:
            continue
        team = str(r["HomeTeam"]).strip()
        if team:
            idx.setdefault((y, w, d), []).append((tmin, team))
    return idx

def fill_matches(
    matrix: pd.DataFrame,
    match_index,
    week_label_style: str,
    slots: Dict[str, List[Tuple[str,str]]],
    activities_index: Optional[Dict[tuple, List[Tuple[int, str]]]] = None
):
    cols = list(matrix.columns)
    
    if activities_index is None:
        activities_index = {}

    for (d, van, tot, regel) in matrix.index:
        if not (van and tot and regel == "Wedstrijden"):
            continue

        v_from = _hhmm_to_minutes(van)
        v_to   = _hhmm_to_minutes(tot)

        if v_from < 0 or v_to <= v_from:
            continue

        for label in cols:

            if week_label_style == "iso":
                parts = label.split("-W")
                y, w = int(parts[0]), int(parts[1])
            else:
                try:
                    w = int(label.split()[1])
                except Exception:
                    continue
                y = now_naive_in_tz(TZ).isocalendar().year

            matches = match_index.get((y, w, d), [])
            activities = activities_index.get((y, w, d), [])

            if not matches and not activities:
                continue
                
            # 🔹 groepeer per tijd
            grouped = {}

            for tmin, team in matches:
                if v_from <= tmin < v_to:
                    grouped.setdefault(tmin, []).append(team)
            
            for tmin, act in activities:
                if v_from <= tmin < v_to:
                    grouped.setdefault(tmin, []).append(act)
            
            if not grouped:
                continue
            
            # 🔹 sorteer tijden
            lines = []
            for tmin in sorted(grouped.keys()):
                hh = tmin // 60
                mm = tmin % 60
                tijd = f"{hh:02d}:{mm:02d}"
                teams = ", ".join(grouped[tmin])
                lines.append(f"{tijd}: {teams}")

            text = "\n".join(lines)

            key = (d, van, tot, "Wedstrijden")
            if key in matrix.index:
                matrix.loc[key, label] = text

def prune_empty_subrows(matrix: pd.DataFrame) -> pd.DataFrame:
    """Verwijder lege subregels, behalve:
       - dag-headers
       - 'Namen'-regels (altijd tonen, ook als leeg) om lege tijdvakken zichtbaar te houden.
    """
    keep_flags = []
    for idx in matrix.index:
        d, van, tot, regel = idx
        # Dag-header altijd houden
        if not van and not tot and not regel:
            keep_flags.append(True)
            continue
        # 'Namen' nooit verwijderen (ook leeg zichtbaar)
        if regel == "Namen":
            keep_flags.append(True)
            continue
        # Overige subregels alleen houden als er inhoud is
        row = matrix.loc[idx]
        has_content = any(bool(str(v)) for v in row.values)
        keep_flags.append(has_content)
    return matrix[keep_flags]

# ===== formatter (rijkere opmaak) =====
def format_sheet(ws, matrix: pd.DataFrame, slots: Dict[str, List[Tuple[str,str]]], tz_str: str):
    thin  = Side(style="thin",  color="FFAAAAAA")
    thick = Side(style="thick", color="FF000000")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    first_week_col_idx = 4  # A:Dag, B:Tijd-van, C:Tijd-tot, D: 1e weekkolom

    # Laatste rij per dag → dikke horizontale lijn
    day_last_row = {}
    for r_idx, (d, van, tot, regel) in enumerate(matrix.index, start=2):
        if van and regel in ("Handmatig","Wedstrijden","Namen"):
            day_last_row[d] = r_idx

    # Opmaak per cel
    for r_idx, (d, van, tot, regel) in enumerate(matrix.index, start=2):
        is_header = (van == "" and tot == "" and regel == "")
        for c_idx in range(1, ws.max_column+1):
            cell = ws.cell(row=r_idx, column=c_idx)
            # dunne randen overal
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            if is_header:
                fill = PatternFill(start_color=DAY_COLORS.get(d, "FFFFFFFF"),
                                   end_color=DAY_COLORS.get(d, "FFFFFFFF"),
                                   fill_type="solid")
                cell.fill = fill
                cell.font = bold
                cell.alignment = center
            else:
                # kolommen B/C: tijden altijd zwart
                if c_idx in (2, 3):
                    cell.font = Font(color="FF000000")
                    cell.alignment = wrap
                else:
                    # weekkolommen: kleur per subregel + wrap
                    if regel in ("Handmatig","Wedstrijden"):
                        cell.font = Font(color="FFCC0000")
                    else:
                        cell.font = Font(color="FF000000")
                    cell.alignment = wrap

        # dikke onderrand op dag-einde
        if r_idx in day_last_row.values():
            for c_idx in range(1, ws.max_column+1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = Border(left=cell.border.left, right=cell.border.right,
                                     top=cell.border.top, bottom=thick)

    # Dikke verticale scheiding tussen weekkolommen
    for j in range(first_week_col_idx, ws.max_column+1):
        for r in range(1, ws.max_row+1):
            cell = ws.cell(row=r, column=j)
            cell.border = Border(left=thick, right=cell.border.right,
                                 top=cell.border.top, bottom=cell.border.bottom)

    # Kolombreedtes (met eenvoudige autosize)
    ws.column_dimensions['A'].width = 12  # Dag
    ws.column_dimensions['B'].width = 9   # Tijd-van
    ws.column_dimensions['C'].width = 9   # Tijd-tot
    for col_cells in ws.iter_cols(min_col=first_week_col_idx, max_col=ws.max_column):
        max_len = 14
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(int(max_len*0.75)+2, 10), 24)

    # Timestamp in A1
    now = now_naive_in_tz(tz_str)
    stamp = f"{now.day} {month_short_nl(now.month)} {now.strftime('%H:%M')}"
    a1 = ws.cell(row=1, column=1); a1.value = stamp
    try: a1.font = Font(italic=True, color="FF666666")
    except Exception: pass

    ws.freeze_panes = "D2"  # tot en met Tijd-tot + header

# ===== Wedstrijden normaliseren =====
def _strip_ckc_prefix(name: str) -> str:
    if not isinstance(name, str):
        return ""
    s = name.strip(); up = s.upper()
    if up.startswith("CKC JO") or up.startswith("CKC MO") or up.startswith("CKC O") or up.startswith("CKC VR"):
        return s[4:].lstrip()
    return s

def normalize_program(data, tz_str: str) -> pd.DataFrame:
    df = pd.DataFrame(data)
    empty = pd.DataFrame(columns=["Datum","Dag","Tijd","ISO_Year","Week","HomeTeam"])
    if df.empty:
        return empty

    c_date = _pick(df.columns, ["wedstrijddatum"])
    c_home = _pick(df.columns, ["thuisteam"])
    c_home_code = _pick(df.columns, ["thuisteamclubrelatiecode"])
    for col in (c_date, c_home, c_home_code):
        if col is None or col not in df.columns:
            return empty

    dt = pd.to_datetime(df[c_date].astype(str).str.strip(), errors="coerce", utc=True)
    mask_ok = dt.notna()
    if not mask_ok.any():
        return empty
    dt_local_naive = dt.dt.tz_convert(ZoneInfo(tz_str)).dt.tz_localize(None)

    df = df.loc[mask_ok].copy()
    df["Datum"] = dt_local_naive.astype("datetime64[ns]")

    df = df[df[c_home_code].astype(str) == CKC_CLUBRELATIECODE].copy()
    if df.empty:
        return empty

    df["Dag"] = df["Datum"].dt.weekday.map(lambda i: DAYS_NL[i])
    df["Tijd"] = df["Datum"].dt.strftime("%H:%M")
    iso = df["Datum"].dt.isocalendar()
    df["ISO_Year"] = iso.year.astype(int)
    df["Week"] = iso.week.astype(int)
    df["HomeTeam"] = df[c_home].astype(str).map(_strip_ckc_prefix)
    return df[["Datum","Dag","Tijd","ISO_Year","Week","HomeTeam"]]

# ====== normaliseer activiteiten ====
def normalize_activities(data, tz_str: str) -> pd.DataFrame:
    df = pd.DataFrame(data)

    empty = pd.DataFrame(columns=["Datum","Dag","Tijd","ISO_Year","Week","Activiteit","Date","IsAllDay"])
    if df.empty:
        return empty

    c_dt = _pick(df.columns, ["datumvan"])
    c_name = _pick(df.columns, ["activiteit"])
    c_heledag = _pick(df.columns, ["heledag"])

    if not c_dt or not c_name:
        return empty

    dt = pd.to_datetime(df[c_dt], errors="coerce", utc=True)
    dt = dt.dt.tz_convert(ZoneInfo(tz_str)).dt.tz_localize(None)

    df["Datum"] = dt
    df = df.dropna(subset=["Datum"])

    # 🔹 BESTAANDE LOGICA (ongewijzigd)
    df["Dag"] = df["Datum"].dt.weekday.map(lambda i: DAYS_NL[i])
    df["Tijd"] = df["Datum"].dt.strftime("%H:%M")

    iso = df["Datum"].dt.isocalendar()
    df["ISO_Year"] = iso.year.astype(int)
    df["Week"] = iso.week.astype(int)

    df["Activiteit"] = df[c_name].astype(str).str.strip()

    # 🔹 hele dag events
    df["IsAllDay"] = False
    if c_heledag:
        mask = df[c_heledag].astype(str).str.lower() == "true"
        df.loc[mask, "Tijd"] = "00:00"
        df.loc[mask, "IsAllDay"] = True

    # 🆕 NIEUW voor kalender-grid
    df["Date"] = df["Datum"].dt.date

    return df[["Datum","Dag","Tijd","ISO_Year","Week","Activiteit","Date","IsAllDay"]]

def build_activities_index(df: pd.DataFrame) -> Dict[tuple, List[Tuple[int, str]]]:
    idx: Dict[tuple, List[Tuple[int, str]]] = {}

    for _, r in df.iterrows():
        y, w, d = int(r["ISO_Year"]), int(r["Week"]), r["Dag"]

        try:
            h, m = r["Tijd"].split(":")
            tmin = int(h)*60 + int(m)
        except Exception:
            continue

        txt = r["Activiteit"]

        if txt:
            idx.setdefault((y, w, d), []).append((tmin, txt))

    return idx

# ===== Handmatige input (.txt) =====
def parse_manual_text(text: str):
    entries = []
    if not text:
        return entries
    now_aw = now_aware_in_tz(TZ)
    monday_aw = (now_aw - pd.Timedelta(days=int(now_aw.weekday()))).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("#"): continue
        parts = s.split()
        if len(parts) < 3: continue
        date_str, time_str = parts[0], parts[1]
        txt = " ".join(parts[2:]).strip()
        try:
            dt_aw = pd.Timestamp(f"{date_str} {time_str}", tz=ZoneInfo(TZ))
        except Exception:
            continue
        if dt_aw < monday_aw: continue
        day_name = DAYS_NL[int(dt_aw.weekday())]
        starts = [a for a, b in DEFAULT_SLOTS.get(day_name, [])]
        if time_str not in starts: continue
        iso = dt_aw.isocalendar()
        entries.append({
            "date": dt_aw.tz_convert(None),
            "time_from": time_str,
            "text": txt,
            "iso_year": int(iso.year),
            "iso_week": int(iso.week),
            "day": day_name,
        })
    return entries

# -------- Dropbox helper --------
def _ensure_dropbox_direct(url: str) -> str:
    if not url or "dropbox.com" not in url:
        return url
    try:
        pr = urlparse(url)
        qs = parse_qs(pr.query); qs["dl"] = ["1"]
        new_query = urlencode({k: v[0] for k, v in qs.items()})
        return urlunparse((pr.scheme, pr.netloc, pr.path, pr.params, new_query, pr.fragment))
    except Exception:
        return url.replace("dl=0", "dl=1") if "dl=" in url else (url + ("&" if "?" in url else "?") + "dl=1")

def read_manual_text_from_dropbox(timeout: int = 30) -> str:
    direct = _ensure_dropbox_direct(DROPBOX_INPUT_URL)
    r = requests.get(direct, timeout=timeout)
    r.raise_for_status()
    return r.text

# ===== Excel bouwen =====
def make_excel(df_bar, df_ck, annotations,
               use_matches=True,
               use_overrides=True,
               use_activities=True, 
                add_activities_sheet=True):    
    
    df_activities = pd.DataFrame()
    
    

    if use_activities or add_activities_sheet:
        activities_url = build_activities_url(ACTIVITIES_DAYS_AHEAD, DEFAULT_CLIENT_ID)
        activities_json = http_get_json(activities_url)
    
        df_activities = normalize_activities(activities_json, TZ)
        df_activities = filter_from_current_week(df_activities, TZ)
        
        
        # 🔹 NIEUW: custom slots integreren
        merged_slots, slot_warnings = merge_custom_slots_into_defaults(
            [df_bar, df_ck],
            DEFAULT_SLOTS,
            df_activities if use_activities else None
        )
        
        activities_index = build_activities_index(df_activities) if not df_activities.empty else {}    
        
        # Weekrange bepalen: uit data + huidige week + annotaties
        def compute_weeks(df_list):
            pairs = set()
            for df in df_list:
                if df is None or df.empty: continue
                for y, w in zip(df["ISO_Year"], df["Week"]):
                    pairs.add((int(y), int(w)))
            now = now_naive_in_tz(TZ); iso = now.isocalendar()
            pairs.add((int(iso.year), int(iso.week)))
            pairs |= {(a["iso_year"], a["iso_week"]) for a in annotations}
            pairs = sorted(pairs)
            wmondays = {p: pd.Timestamp.fromisocalendar(p[0], p[1], 1) for p in pairs}
            return pairs, wmondays
    
        weeks_pairs, week_mondays = compute_weeks([df_bar, df_ck])
    
        # Lege matrixen opzetten
        def build_empty(slots, subset):
            return build_empty_matrix(slots, TZ, subset, 4, WEEK_LABEL, weeks_pairs, week_mondays)
    
        days_subset_ck = ["Zaterdag"] if SAT_ONLY_CK else None
        matrix_bar = build_empty(merged_slots, None)
        matrix_ck  = build_empty(merged_slots, days_subset_ck)
    
        # Vullen: namen (overlap + warnings)
        warn_bar = fill_names(matrix_bar, df_bar, merged_slots, WEEK_LABEL)
        warn_ck  = fill_names(matrix_ck,  df_ck,  merged_slots, WEEK_LABEL)
        warnings_total = warn_bar + warn_ck + slot_warnings
    
        # Overrides toepassen
        if use_overrides:
            overrides = load_afgeschermd_overrides_from_dropbox(debug_fetch)
        
            apply_afgeschermd_overrides(matrix_bar, overrides, "bar", WEEK_LABEL, debug_fetch)
            apply_afgeschermd_overrides(matrix_ck,  overrides, "ck",  WEEK_LABEL, debug_fetch)
    
        # Vullen: handmatig
        fill_manual(matrix_bar, annotations, merged_slots, WEEK_LABEL)
        fill_manual(matrix_ck,  annotations, merged_slots, WEEK_LABEL)
    
        # Vullen: wedstrijden
        if use_matches:
            program_url = build_program_url(PROGRAM_DAYS_AHEAD, DEFAULT_CLIENT_ID, PROGRAM_FIELDS,
                                            eigenwedstrijden="JA", thuis="JA", uit="NEE",
                                            gebruiklokaleteamgegevens="NEE")
            program_json = http_get_json(program_url)
            df_program = normalize_program(program_json, TZ)
            df_program = filter_from_current_week(df_program, TZ)
            match_index = build_match_index_for_overlap(df_program)
            fill_matches(matrix_bar, match_index, WEEK_LABEL, merged_slots, activities_index)
            fill_matches(matrix_ck,  match_index, WEEK_LABEL, merged_slots, activities_index)
        
        
        
        
        # Prune: verwijder subregels die volledig leeg zijn, maar laat 'Namen' altijd staan
        matrix_bar = prune_empty_subrows(matrix_bar)
        matrix_ck  = prune_empty_subrows(matrix_ck)
    
        matrix_activities_calendar = None
    
        if add_activities_sheet and not df_activities.empty:
            matrix_activities_calendar = build_activities_calendar_matrix(
                df_activities
)
    
        # Schrijf naar Excel en verwijder de 'Regel'-kolom (kolom D)
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            # BarRooster
            matrix_bar.to_excel(writer, sheet_name="BarRooster")
            ws_bar = writer.sheets["BarRooster"]
            ws_bar.delete_cols(4)  # kolom D
            format_sheet(ws_bar, matrix_bar, merged_slots, TZ)
    
            # CommissieKamer
            matrix_ck.to_excel(writer, sheet_name="CommissieKamer")
            ws_ck = writer.sheets["CommissieKamer"]
            ws_ck.delete_cols(4)
            format_sheet(ws_ck, matrix_ck, merged_slots, TZ)
            
            # 🆕 Activiteiten Kalender
            if matrix_activities_calendar is not None:
    
                sheet_name = "Activiteiten"
            
                matrix_activities_calendar.to_excel(writer, sheet_name=sheet_name)
            
                ws_act = writer.sheets[sheet_name]
            
                format_activities_calendar_sheet(ws_act, matrix_activities_calendar, TZ)
                    
            bio.seek(0)
        return bio, warnings_total

# =========================
# UI (simpel)
# =========================
st.set_page_config(page_title=f"CKC Rooster generator v{__version__}", page_icon="🗓️", layout="centered")
st.markdown("<h1 style='text-align:center;margin-bottom:0'>CKC Rooster generator</h1>", unsafe_allow_html=True)
st.markdown(f"<h5 style='text-align:center;margin-top:0.25rem;color:#666'>versie {__version__}</h5>", unsafe_allow_html=True)
st.caption("Sportlink → Excel · vaste instellingen (Europe/Amsterdam), weekoffset=-1, gefilterd vanaf huidige week")

add_activities_sheet = st.checkbox(
    "Toon activiteiten kalender",
    value=True
)
use_dropbox = st.checkbox("Handmatige input via Dropbox meenemen", value=True)
use_matches = st.checkbox("Wedstrijdinfo toevoegen", value=True)
use_overrides = st.checkbox("Gebruik Afgeschermd overrides", value=True)
use_activities = st.checkbox("Verenigingsagenda toevoegen", value=True)

# debug_fetch = st.checkbox("Toon Sportlink fetch logging", value=False)

if st.checkbox("Toon Sportlink fetch logging (debug modus)", key="debug_fetch"):
    st.cache_data.clear()
    debug_fetch = True
    st.info("Debug modus actief (cache uitgeschakeld)")
else:
    debug_fetch = False

if st.button("Genereer rooster", use_container_width=True):
    try:
        with st.spinner("Ophalen en bouwen…"):
            # Vrijwilligersdata
            urls_bar = build_urls(BAR_CODES, DAYS_AHEAD, DEFAULT_CLIENT_ID, weekoffset=WEEK_OFFSET, fields=FIELDS)
            urls_ck  = build_urls(CK_CODES,  DAYS_AHEAD, DEFAULT_CLIENT_ID, weekoffset=WEEK_OFFSET, fields=FIELDS)
                        
            all_bar = fetch_all(urls_bar, debug_fetch)
            all_ck  = fetch_all(urls_ck, debug_fetch)
            
            df_bar = filter_from_current_week(normalize_dataframe(all_bar, TZ), TZ)
            df_ck  = filter_from_current_week(normalize_dataframe(all_ck,  TZ), TZ)

            # Handmatige input
            manual_text = ""
            if use_dropbox:
                direct = _ensure_dropbox_direct(DROPBOX_INPUT_URL)
                
                try:
                    manual_text = session.get(direct, timeout=30).text
                except Exception:
                    manual_text = ""
                    st.warning("Dropbox handmatige input kon niet worden opgehaald.")                   
            
            annotations = parse_manual_text(manual_text)
                
            # Excel bouwen (met/zonder wedstrijden) + waarschuwingen
            xlsx, warnings_total = make_excel(
                df_bar,
                df_ck,
                annotations,
                use_matches=use_matches,
                use_overrides=use_overrides,
                use_activities=use_activities,
                add_activities_sheet=add_activities_sheet
            )



        # Waarschuwingen tonen (als aanwezig)
        slot_msgs = [w for w in warnings_total if "slot" in w.lower()]
        placement_msgs = [w for w in warnings_total if w not in slot_msgs]

        if slot_msgs:
            st.info("ℹ️ Tijdsloten automatisch toegevoegd:\n\n- " + "\n- ".join(slot_msgs))
        
        if placement_msgs:
            st.warning("⚠️ Niet alle diensten konden worden geplaatst:\n\n- " + "\n- ".join(placement_msgs))            
        
        
        calls = sportlink_stats["calls"]
                
        retries = sportlink_stats["retries"]
        failures = sportlink_stats["failures"]

        if failures > 0:
            st.error(f"Sportlink status: probleem ({failures} fetch mislukt)")
        elif retries > 0:
            st.warning(f"Sportlink status: instabiel ({retries} retries nodig)")
        else:
            st.success("Sportlink status: OK")

        st.success("Klaar! Download hieronder het Excel-bestand.")
        st.download_button(
            "⬇️ Download rooster.xlsx",
            data=xlsx.getvalue(),
            file_name="rooster.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except requests.HTTPError as e:
        st.error(f"Fout bij ophalen gegevens: {e}")
    except Exception as e:
        st.error(f"Er ging iets mis: {e}")